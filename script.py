import os
import re
import openpyxl
from openpyxl import Workbook

pasta_fichas = r"Caminho\pastas"

# Criação da planilha geral
wb_geral = Workbook()
ws_geral = wb_geral.active
ws_geral.title = "Reeducandos"
ws_geral.append(["Nome", "CPF", "Telefone", "Frequência", "Arquivo Origem"])

for nome_arquivo in os.listdir(pasta_fichas):
    if nome_arquivo.endswith(".xlsx") and not nome_arquivo.startswith("~$"):
        caminho = os.path.join(pasta_fichas, nome_arquivo)
        try:
            wb = openpyxl.load_workbook(caminho, data_only=True)
            ws = wb.active

            # Acessa diretamente as células com base no modelo padrão
            nome = ws["B4"].value
            cpf = ws["G5"].value
            telefone = ws["B8"].value

            # Frequência
            frequencia = "Não informado"
            freq_raw = ws["A2"].value
            if freq_raw:
                freq_upper = str(freq_raw).upper()

                if re.search(r"\(\s*X\s*\)\s*MENSAL", freq_upper):
                    frequencia = "Mensal"
                elif re.search(r"\(\s*X\s*\)\s*BIMESTRAL", freq_upper):
                    frequencia = "Bimestral"
                elif re.search(r"\(\s*X\s*\)\s*OUTROS", freq_upper):
                    frequencia = "Outro"

            ws_geral.append([nome, cpf, telefone, frequencia, nome_arquivo])

        except Exception as e:
            print(f"Erro ao processar {nome_arquivo}: {e}")

# Salva a planilha final
wb_geral.save("planilha_reeducandos.xlsx")
print("✅ Planilha geral criada com sucesso.")
